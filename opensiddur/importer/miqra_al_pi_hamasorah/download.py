"""Download Miqra al pi ha-Masorah from Google Sheets into per-tab TSV files."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

from opensiddur.importer.util.pages import (
    default_sourcetexts_root,
    miqra_al_pi_hamasorah_data_directory,
    miqra_al_pi_hamasorah_sheets_directory,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

SPREADSHEET_ID = "1mkQyj6by1AtBUabpbaxaZq9Z2X3pX8ZpwG91ZCSOEYs"
SOURCE_URL = (
    f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/edit"
)
EXPORT_XLSX_URL = (
    f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
)
USER_AGENT = (
    "OpenSiddur-AI/1.0 (https://github.com/opensiddur/opensiddur-ai; "
    "opensiddur@example.com)"
)

# Exact worksheet titles from the workbook → output slug (without .tsv).
SHEET_SLUGS: dict[str, str] = {
    "שינויים changes": "changes",
    "README": "readme",
    "כתובים אחרונים": "ketuvim_aharonim",
    "חמש מגילות": "chamisha_megillot",
    "ספרי אמ\"ת": "sifrei_emet",
    "נביאים אחרונים": "neviim_acharonim",
    "נביאים ראשונים": "neviim_rishonim",
    "תורה": "torah",
    "תבניות templates": "templates",
    "מיוחד special": "special",
    "AutoEdits": "auto_edits",
}


def _cell_value(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _worksheet_rows(worksheet: Any) -> tuple[list[list[str]], int, int]:
    """Return (rows, row_count, max_columns) for a worksheet."""
    rows: list[list[str]] = []
    max_col = 0
    for row in worksheet.iter_rows(values_only=True):
        cells = [_cell_value(c) for c in row]
        while cells and cells[-1] == "":
            cells.pop()
        if not any(cells):
            continue
        max_col = max(max_col, len(cells))
        rows.append(cells)
    if max_col == 0:
        return [], 0, 0
    padded = [cells + [""] * (max_col - len(cells)) for cells in rows]
    return padded, len(padded), max_col


def _write_tsv(path: Path, rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(
            f,
            delimiter="\t",
            lineterminator="\n",
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writerows(rows)


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _split_workbook(xlsx_path: Path, sheets_dir: Path) -> list[dict[str, Any]]:
    sheet_entries: list[dict[str, Any]] = []
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            title = worksheet.title
            slug = SHEET_SLUGS.get(title)
            if slug is None:
                logger.warning("Skipping unknown worksheet: %r", title)
                continue
            rows, row_count, col_count = _worksheet_rows(worksheet)
            out_path = sheets_dir / f"{slug}.tsv"
            _write_tsv(out_path, rows)
            rel_path = f"sheets/{slug}.tsv"
            sheet_entries.append(
                {
                    "name": title,
                    "slug": slug,
                    "path": rel_path,
                    "rows": row_count,
                    "columns": col_count,
                }
            )
            logger.info("Wrote %s (%d rows, %d columns)", out_path, row_count, col_count)
    finally:
        workbook.close()
    return sheet_entries


def download_miqra(
    sourcetexts_root: Path | None = None,
    *,
    dry_run: bool = False,
) -> None:
    """Download the spreadsheet and write per-tab TSV files plus manifest.json."""
    data_dir = miqra_al_pi_hamasorah_data_directory(sourcetexts_root)
    sheets_dir = miqra_al_pi_hamasorah_sheets_directory(sourcetexts_root)
    manifest_path = data_dir / "manifest.json"

    if dry_run:
        logger.info("Would download %s", EXPORT_XLSX_URL)
        logger.info("Would write TSV files under %s", sheets_dir)
        logger.info("Would write manifest to %s", manifest_path)
        return

    data_dir.mkdir(parents=True, exist_ok=True)
    sheets_dir.mkdir(parents=True, exist_ok=True)

    headers = {"User-Agent": USER_AGENT}
    logger.info("Downloading %s ...", EXPORT_XLSX_URL)
    response = requests.get(EXPORT_XLSX_URL, headers=headers, timeout=300)
    response.raise_for_status()

    tmp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            suffix=".xlsx",
            delete=False,
            dir=data_dir,
        ) as tmp:
            tmp.write(response.content)
            tmp_path = Path(tmp.name)

        logger.info("Splitting workbook into TSV files ...")
        sheet_entries = _split_workbook(tmp_path, sheets_dir)

        for entry in sheet_entries:
            tsv_path = data_dir / entry["path"]
            entry["sha256"] = _sha256_file(tsv_path)

        manifest = {
            "spreadsheet_id": SPREADSHEET_ID,
            "source_url": SOURCE_URL,
            "export_url": EXPORT_XLSX_URL,
            "downloaded_at": datetime.now(timezone.utc).isoformat(),
            "sheets": sheet_entries,
        }
        manifest_path.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        logger.info("Wrote manifest to %s", manifest_path)
    finally:
        if tmp_path is not None and tmp_path.exists():
            tmp_path.unlink()
            logger.info("Removed temporary workbook %s", tmp_path)


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Download Miqra al pi ha-Masorah from Google Sheets into per-tab TSV "
            "files under <sourcetexts-root>/miqra_al_pi_hamasorah."
        )
    )
    parser.add_argument(
        "--sourcetexts-root",
        type=Path,
        default=default_sourcetexts_root(),
        help=(
            "Root of the sourcetexts tree; output is written under "
            "<root>/miqra_al_pi_hamasorah (default: <repo>/sources)."
        ),
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Log actions without downloading or writing files.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = _build_arg_parser().parse_args(argv)
    download_miqra(args.sourcetexts_root, dry_run=args.dry_run)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as e:
        logger.error("Error downloading Miqra al pi ha-Masorah: %s", e)
        raise
